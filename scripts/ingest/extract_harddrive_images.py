#!/usr/bin/env python3
"""
Extract Image URLs from HardDrive Excel File
Extracts real URLs from HYPERLINK formulas in Excel
"""

import openpyxl
import csv
import sys
import re

def extract_url_from_formula(cell_value):
    """Extract URL from Excel HYPERLINK formula"""
    if cell_value is None:
        return None
    
    # Convert to string
    formula_str = str(cell_value)
    
    # Match HYPERLINK formula: =HYPERLINK("url", "display")
    # or just the URL pattern directly
    url_pattern = r'(https?://[^\s"]+\.(?:jpg|jpeg|png|gif|webp))'
    match = re.search(url_pattern, formula_str, re.IGNORECASE)
    
    if match:
        return match.group(1)
    
    return None

def process_excel_file(input_file, output_file):
    """Process Excel file and extract image URLs"""
    print(f"📖 Reading Excel file: {input_file}")
    
    try:
        wb = openpyxl.load_workbook(input_file, data_only=False)
        ws = wb.active
        
        print(f"   Worksheet: {ws.title}")
        print(f"   Rows: {ws.max_row:,}")
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"\n📋 Headers found: {', '.join(headers)}\n")
        
        # Find image_uri column
        image_col_idx = None
        for idx, header in enumerate(headers):
            if header and 'image' in header.lower() and 'uri' in header.lower():
                image_col_idx = idx
                break
        
        if image_col_idx is None:
            print("⚠️  Could not find 'image_uri' column")
            print("   Available columns:", headers)
            return
        
        print(f"✓ Found image_uri column at index {image_col_idx}\n")
        
        # Process all rows
        rows_processed = 0
        urls_found = 0
        
        output_data = []
        output_data.append(headers)  # Add header row
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            row_data = []
            
            for col_idx, cell in enumerate(row):
                if col_idx == image_col_idx:
                    # Extract URL from formula
                    url = extract_url_from_formula(cell.value)
                    row_data.append(url or '')
                    
                    if url:
                        urls_found += 1
                else:
                    # Regular cell value
                    row_data.append(cell.value)
            
            output_data.append(row_data)
            rows_processed += 1
            
            # Progress indicator
            if rows_processed % 5000 == 0:
                print(f"   Processed {rows_processed:,} rows, found {urls_found:,} URLs...")
        
        print(f"\n✅ Processing complete:")
        print(f"   Rows processed: {rows_processed:,}")
        print(f"   Image URLs found: {urls_found:,}")
        
        # Write to CSV
        print(f"\n💾 Writing CSV: {output_file}")
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in output_data:
                writer.writerow(row)
        
        print(f"✅ CSV written: {output_file}")
        
        # Sample a few URLs
        print("\n📸 Sample image URLs:")
        sample_count = 0
        for row in output_data[1:]:  # Skip header
            if row[image_col_idx] and sample_count < 5:
                print(f"   {row[0]}: {row[image_col_idx]}")
                sample_count += 1
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("""
HardDrive Image URL Extractor

Usage:
  python3 extract_harddrive_images.py <input.xlsx> [output.csv]

Example:
  python3 extract_harddrive_images.py Harddrive-Image-List-0426.xlsx

This will:
  1. Read the Excel file with HYPERLINK formulas
  2. Extract real image URLs from formulas
  3. Output a clean CSV with actual URLs
        """)
        sys.exit(0)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.replace('.xlsx', '_with_urls.csv')
    
    process_excel_file(input_file, output_file)
    
    print(f"\n🎉 Done! Import with:")
    print(f"   node scripts/ingest/import_harddrive_imagelist.js {output_file}\n")
