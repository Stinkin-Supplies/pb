#!/usr/bin/env python3
import openpyxl
import csv
import re
from pathlib import Path

def extract_hyperlinks(excel_file, output_csv):
    print(f"Opening {excel_file}...")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    print(f"Found {ws.max_row} rows")
    
    headers = [cell.value for cell in ws[1]]
    sku_col = headers.index('sku') if 'sku' in headers else 0
    image_col = headers.index('primary_item_image') if 'primary_item_image' in headers else -1
    
    print(f"SKU column: {sku_col}, Image column: {image_col}")
    
    with open(output_csv, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['sku', 'image_url'])
        
        count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            sku = row[sku_col].value
            image_cell = row[image_col]
            
            # Extract URL from formula text
            image_url = None
            if image_cell.value:
                # Match URL inside HYPERLINK formula
                match = re.search(r'"(http[^"]+)"', str(image_cell.value))
                if match:
                    image_url = match.group(1)
            
            if sku and image_url:
                writer.writerow([sku, image_url])
                count += 1
                
            if row_idx % 1000 == 0:
                print(f"  Processed {row_idx} rows, extracted {count} images...")
    
    print(f"\n✓ Done! Extracted {count} images to {output_csv}")

if __name__ == '__main__':
    excel_path = Path('scripts/data/wps/harddrive_master_image.xlsx')
    csv_path = Path('scripts/data/wps/wps_hd_images.csv')
    
    if not excel_path.exists():
        print(f"ERROR: {excel_path} not found")
        exit(1)
    
    extract_hyperlinks(excel_path, csv_path)
